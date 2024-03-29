from django.shortcuts import render, redirect
from django.urls import reverse
from django.contrib import messages
from openpyxl import load_workbook
from django.core.paginator import Paginator
from .models import *
from .forms import *
from django.views.generic.list import ListView
from django.views.generic.detail import DetailView
from django.views import View
from django.contrib import messages
import uuid
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth.decorators import login_required
from django.contrib.auth import authenticate, login, logout

def custom_404_view(request, exception):
    return render(request, '404.html', status=404)


def error_500(request):
        data = {}
        return render(request,'myapp/error_500.html', data)

def upload_excel(request):
    if request.method == 'POST':
        excel_file = request.FILES.get('excel_file')

        if not excel_file.name.endswith('.xlsx'):
            messages.error(request, 'Please upload a valid Excel file (.xlsx)')
        else:
            try:
                workbook = load_workbook(excel_file)
                sheet = workbook.active

                for row in sheet.iter_rows(min_row=2, values_only=True):
                    if len(row) >= 8:
                        Shahar, Mahallasi, Manzil, Ijarachi, Telefon, Yigit, Qiz, Holati, *extra = row
                        
                        # Set empty cells to None
                        if not any(cell for cell in row):
                            Shahar = Mahallasi = Manzil = Ijarachi = Telefon = Yigit = Qiz = Holati = None
                        
                        Home.objects.create(
                            Shahar=Shahar, Mahallasi=Mahallasi, Manzil=Manzil,
                            Ijarachi=Ijarachi, Telefon=Telefon, Yigit=Yigit,
                            Qiz=Qiz, Holati=Holati
                        )
                    else:
                        messages.error(request, 'Row does not contain enough values')

                messages.success(request, 'Data from Excel file imported successfully')
            except Exception as e:
                messages.error(request, f'An error occurred: {str(e)}')

    return render(request, 'upload_excel.html')

# class Home(ListView):
#     paginate_by = 25
#     template_name = 'index.html'
#     model = Home
#     context_object_name = 'ijara' 


class AboutListView(ListView):
    paginate_by = 3
    template_name = 'about.html'
    model = Home
    context_object_name = 'ijara'

class HomeListView(ListView):
    # paginate_by = 15
    template_name = 'index.html'
    model = Ariza
    context_object_name = 'ijara'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['form'] = ArizaForm()

        context['yotoqxonalar'] = Yotoqxonalar.objects.all() 
        return context
    
class ArizaDetailView(DetailView):
    model = Ariza
    template_name = 'ariza_detail.html'
    context_object_name = 'ariza'
    slug_field = 'unique_id'
    slug_url_kwarg = 'unique_id'

    def get_queryset(self):
        queryset = super().get_queryset()
        if not self.request.user.is_staff:
            queryset = queryset.filter(user=self.request.user)
        return queryset

def user_login(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(username=username, password=password)
        if user is not None:
            login(request, user)
            return redirect('ariza_list')  # Redirect to dashboard view after successful login
        else:
            error_message = "Invalid login credentials. Please try again."
            return render(request, 'registration/login.html', {'error_message': error_message})
    return render(request, 'registration/login.html')

@login_required 
def custom_logout(request):
    logout(request)
    return redirect(reverse('login'))  

@login_required
def ariza_list(request):
    if request.method == 'POST':
        action = request.POST.get('action')
        ariza_ids = request.POST.getlist('ariza_ids')
        
        if action == 'approve':
            Ariza.objects.filter(id__in=ariza_ids).update(approved=True)
        elif action == 'reject':
            Ariza.objects.filter(id__in=ariza_ids).update(approved=False)
    arizalar = Ariza.objects.all().order_by('-approved', 'ism')
    
    context = {
        'arizalar': arizalar,
    }
    
    return render(request, 'ariza_list.html', context)
class ArizaDetailView(DetailView):
    model = Ariza
    template_name = 'ariza_detail.html'
    context_object_name = 'ariza'
    slug_field = 'unique_id'
    slug_url_kwarg = 'unique_id'

    def get_queryset(self):
        queryset = super().get_queryset()
        if not self.request.user.is_staff:
            queryset = queryset.filter(user=self.request.user)
        return queryset

    def post(self, request, *args, **kwargs):
        ariza = self.get_object()

        if 'approve' in request.POST:
            ariza.approve()
            messages.success(request, f'Application by {ariza.ism} has been approved.')
        elif 'reject' in request.POST:
            ariza.reject()
            messages.success(request, f'Application by {ariza.ism} has been rejected.')
        elif 'add_comment' in request.POST:
            comment = request.POST.get('comment')
            if comment:
                ariza.comments.create(user=request.user, text=comment)
                messages.success(request, 'Comment added successfully.')

        return redirect(reverse('ariza_detail', kwargs={'unique_id': ariza.unique_id}))

ariza_detail_view = ArizaDetailView.as_view()

def create_ariza(request):
    if request.method == 'POST':
        form = ArizaForm(request.POST, request.FILES)
        if form.is_valid():
            unique_id = uuid.uuid4()
            ariza = form.save(commit=False)
            ariza.unique_id = unique_id
            ariza.save()
            messages.success(request, f'Your application has been submitted. Your ID number is: {unique_id}')
            # return redirect('home_list')
        else:
            print(form.errors)
    else:
        form = ArizaForm()
    context = {'form': form}
    return render(request, 'create_ariza.html', context)

#brak kod
def check_status(request):
    if request.method == 'POST':
        unique_id = request.POST.get('unique_id')
        try:
            ariza = Ariza.objects.get(unique_id=unique_id)
            context = {'ariza': ariza}
            return render(request, 'ariza_detail.html', context)
        except Ariza.DoesNotExist:
            message = f'Application with ID {unique_id} not found.'
            context = {'message': message}
            return render(request, 'home.html', context)
    else:
        return redirect('home')
#brak kode

def search_application(request):
    if request.method == 'POST':
        unique_id = request.POST.get('unique_id')
        try:
            ariza = Ariza.objects.get(unique_id=unique_id)
            context = {'ariza': ariza}
            return render(request, 'search_result.html', context)
        except Ariza.DoesNotExist:
            message = f'Application with ID {unique_id} not found.'
            context = {'message': message}
            return render(request, 'search_result.html', context)
    else:
        return render(request, 'search_form.html')

def approve_ariza(request, unique_id):
    ariza = Ariza.objects.get(unique_id=unique_id)
    ariza.approve()
    messages.success(request, f'Application by {ariza.ism} has been approved.')
    return redirect('home_list')

def test(request):
    return render(request, 'test.html',)

def approve_or_reject_application(request, unique_id):
    if request.method == 'POST':
        ariza = Ariza.objects.get(unique_id=unique_id)
        action = request.POST.get('action')  # This should be a hidden input in your template

        if action == 'approve':
            ariza.approve()  # Assuming you have an 'approve' method in your Ariza model
            message = f'Application by {ariza.ism} has been approved.'
        elif action == 'reject':
            ariza.reject()  # You can add a 'reject' method in your Ariza model
            message = f'Application by {ariza.ism} has been rejected.'
        else:
            message = 'Invalid action.'

        context = {'message': message}
        return render(request, 'approval_result.html', context)
    else:
        ariza = Ariza.objects.get(unique_id=unique_id)
        context = {'ariza': ariza}
        return render(request, 'approve_or_reject.html', context)